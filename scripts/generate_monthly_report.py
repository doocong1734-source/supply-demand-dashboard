#!/usr/bin/env python3
"""매월 1일 11:00 - 월간 리포트 생성
- Portfolio Trades에서 지난 달 거래 기록
- track-returns 데이터 분석
- 상세 마크다운 리포트 생성
"""
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os
import glob
import csv

tracker_file = "/Users/kimdoohwan/Documents/Obsidian Vault/주식/스크리너백테스트/매매기록.xlsx"
data_dir = os.path.expanduser("~/Documents/Obsidian Vault/주식/스크리너백테스트")
report_dir = os.path.expanduser("~/Documents/Obsidian Vault/주식/백테스트_리뷰")

try:
    os.makedirs(report_dir, exist_ok=True)

    month = datetime.now().strftime("%Y-%m")
    last_month = (datetime.now().replace(day=1) - timedelta(days=1)).strftime("%Y-%m")

    wb = load_workbook(tracker_file)
    pt_ws = wb["Portfolio Trades"]
    ms_ws = wb["Monthly Summary"]

    # 지난 달 거래 수집
    last_month_trades = []
    for row in range(2, pt_ws.max_row + 1):
        entry_date = pt_ws.cell(row=row, column=6).value
        if not entry_date:
            continue

        entry_date_str = str(entry_date)[:7]
        if entry_date_str == last_month or entry_date_str.startswith(last_month):
            ticker = pt_ws.cell(row=row, column=2).value
            shares = pt_ws.cell(row=row, column=5).value
            entry_price = pt_ws.cell(row=row, column=7).value
            exit_price = pt_ws.cell(row=row, column=8).value

            last_month_trades.append({
                'ticker': ticker,
                'shares': shares,
                'entry_date': entry_date,
                'entry_price': entry_price,
                'exit_price': exit_price,
            })

    # track-returns 분석
    returns_summary = {}
    returns_files = sorted(glob.glob(f"{data_dir}/*_수익률추적.csv"))

    for rf in returns_files:
        try:
            with open(rf, 'r') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    ticker = row.get('ticker', '')
                    if ticker not in returns_summary:
                        returns_summary[ticker] = []
                    returns_summary[ticker].append({
                        'screen_date': row.get('screen_date', ''),
                        'days': int(row.get('days', 0)),
                        'return_pct': float(row.get('return_pct', 0)),
                    })
        except:
            pass

    # 리포트 생성
    lines = [
        f"# {last_month} 성과 분석",
        "",
        f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "## 📊 월간 통계",
        "",
    ]

    # Monthly Summary에서 통계 읽기
    for row in range(2, ms_ws.max_row + 1):
        report_month = ms_ws.cell(row=row, column=1).value
        if report_month and str(report_month).startswith(last_month):
            trades = ms_ws.cell(row=row, column=2).value or 0
            pnl = ms_ws.cell(row=row, column=3).value or 0
            avg_return = ms_ws.cell(row=row, column=4).value or 0

            lines.extend([
                f"- **총 거래**: {int(trades)}개",
                f"- **총 손익**: ₩{float(pnl):,.0f}",
                f"- **평균 수익률**: {float(avg_return):+.2f}%",
                "",
            ])
            break

    # 거래 상세
    if last_month_trades:
        lines.extend([
            "## 📈 거래 상세",
            "",
            "| 종목 | 진입 | 진입가 | 청산가 | 손익 |",
            "|-----|------|--------|--------|------|",
        ])

        for trade in last_month_trades:
            if trade['exit_price']:
                try:
                    pnl = (float(trade['exit_price']) - float(trade['entry_price'])) * float(trade['shares'])
                    return_pct = ((float(trade['exit_price']) - float(trade['entry_price'])) / float(trade['entry_price'])) * 100
                    lines.append(
                        f"| {trade['ticker']} | {trade['entry_date']} | {trade['entry_price']} | {trade['exit_price']} | {return_pct:+.1f}% |"
                    )
                except:
                    pass

        lines.append("")

    # 수익률 추적
    lines.extend([
        "## 📉 스크리너 검증",
        "",
        "### 7일 후 수익률",
        "",
    ])

    for period in [7, 14, 30]:
        period_returns = []
        for ticker, rets in returns_summary.items():
            for ret in rets:
                if ret['days'] == period:
                    period_returns.append((ticker, ret['return_pct']))

        if period_returns:
            avg = sum(r[1] for r in period_returns) / len(period_returns)
            lines.append(f"- **{period}일**: 평균 {avg:+.2f}%")

    lines.extend([
        "",
        "---",
        f"생성: 자동 백테스팅 시스템",
    ])

    report_file = f"{report_dir}/{last_month}_결과.md"
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"✅ 월간 리포트 생성 완료: {report_file}")

except Exception as e:
    print(f"❌ 에러: {e}")
    import traceback
    traceback.print_exc()
