#!/usr/bin/env python3
"""매주 월요일 10:00 - 주간 거래 분석 리포트 생성
- Portfolio Trades에서 지난 주 거래 내역 조회
- 현재 보유 현황 정리
- 손절 모니터링
- 주간 성과 분석 및 마크다운 리포트 생성
"""

from datetime import datetime, timedelta
from openpyxl import load_workbook
import os
import json
import subprocess

tracker_file = "/Users/kimdoohwan/Documents/Obsidian Vault/주식/스크리너백테스트/매매기록.xlsx"
report_dir = os.path.expanduser("~/Documents/Obsidian Vault/주식/백테스트_리뷰")
os.makedirs(report_dir, exist_ok=True)

def get_last_week_range():
    """지난주 범위 반환 (월요일~일요일)"""
    today = datetime.now()
    current_weekday = today.weekday()
    
    last_monday = today - timedelta(days=current_weekday + 7)
    last_sunday = last_monday + timedelta(days=6)
    
    return last_monday.strftime("%Y-%m-%d"), last_sunday.strftime("%Y-%m-%d")

def get_stop_loss_data():
    """손절 모니터링 데이터 조회"""
    try:
        # stop_loss_monitor.py 실행
        subprocess.run([
            'python3',
            '/Users/kimdoohwan/supply-demand-dashboard/scripts/stop_loss_monitor.py'
        ], capture_output=True, timeout=30)
        
        # 결과 파일 읽기
        if os.path.exists('/tmp/stop_loss_data.json'):
            with open('/tmp/stop_loss_data.json', 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"  ⚠️  손절 데이터 조회 실패: {e}")
    
    return None

try:
    wb = load_workbook(tracker_file)
    
    # Portfolio Trades 시트에서 거래 조회
    trades_ws = wb["Portfolio Trades"]
    
    start_date, end_date = get_last_week_range()
    print(f"📊 주간 분석: {start_date} ~ {end_date}")
    
    weekly_trades_buy = []
    weekly_trades_sell = []
    
    for row in range(3, trades_ws.max_row + 1):
        trade_date = trades_ws.cell(row=row, column=2).value
        action = trades_ws.cell(row=row, column=3).value
        ticker = trades_ws.cell(row=row, column=4).value
        company = trades_ws.cell(row=row, column=5).value
        price = trades_ws.cell(row=row, column=6).value
        shares = trades_ws.cell(row=row, column=7).value
        
        if not ticker or not trade_date:
            continue
        
        if isinstance(trade_date, datetime):
            trade_date_str = trade_date.strftime("%Y-%m-%d")
        else:
            trade_date_str = str(trade_date).split(' ')[0]
        
        if start_date <= trade_date_str <= end_date:
            trade_info = {
                'date': trade_date_str,
                'action': action or '',
                'ticker': ticker,
                'company': company or '',
                'price': price or 0,
                'shares': shares or 0,
            }
            
            if action and action.upper() == 'BUY':
                weekly_trades_buy.append(trade_info)
            elif action and action.upper() == 'SELL':
                weekly_trades_sell.append(trade_info)
    
    # Portfolio Status 시트에서 현재 보유 현황
    status_ws = wb["Portfolio Status"]
    current_holdings = []
    total_value = 0
    
    for row in range(2, status_ws.max_row + 1):
        ticker = status_ws.cell(row=row, column=2).value
        shares = status_ws.cell(row=row, column=3).value
        price = status_ws.cell(row=row, column=4).value
        pl = status_ws.cell(row=row, column=5).value
        
        if ticker and shares:
            try:
                shares_num = float(shares) if isinstance(shares, str) else shares
                price_num = float(price) if isinstance(price, str) else (price or 0)
                pl_num = float(pl) if isinstance(pl, str) else (pl or 0)
                
                if shares_num > 0:
                    value = shares_num * price_num
                    total_value += value
                    
                    # 수익률 계산
                    entry_price = status_ws.cell(row=row, column=5).value
                    if entry_price:
                        try:
                            entry_num = float(entry_price)
                            return_rate = (price_num - entry_num) / entry_num if entry_num > 0 else 0
                        except:
                            return_rate = 0
                    else:
                        return_rate = 0
                    
                    current_holdings.append({
                        'ticker': ticker,
                        'shares': shares_num,
                        'price': price_num,
                        'pl': pl_num,
                        'value': value,
                        'return_rate': return_rate
                    })
            except (ValueError, TypeError):
                pass
    
    current_holdings.sort(key=lambda x: x['pl'], reverse=True)
    
    # Daily Screener에서 현재주 Top 10
    ds_ws = wb["Daily Screener"]
    current_week_tickers = []
    
    for row in range(2, ds_ws.max_row + 1):
        date_val = ds_ws.cell(row=row, column=1).value
        ticker = ds_ws.cell(row=row, column=4).value
        company = ds_ws.cell(row=row, column=5).value
        rank = ds_ws.cell(row=row, column=3).value
        
        if ticker and rank and rank <= 10:
            if date_val and str(date_val) == datetime.now().strftime("%Y-%m-%d"):
                current_week_tickers.append((ticker, company))
    
    # 손절 모니터링 데이터 조회
    print("  손절 모니터링 수행 중...")
    stop_loss_data = get_stop_loss_data()
    
    # 마크다운 리포트 생성
    today = datetime.now()
    week_num = today.isocalendar()[1]
    year = today.isocalendar()[0]
    
    report_file = f"{report_dir}/{year}W{week_num:02d}_주간분석.md"
    
    lines = [
        f"# 주간 거래 분석 리포트 - {year}년 {week_num}주차",
        f"",
        f"**기간**: {start_date} ~ {end_date}",
        f"**생성일**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"",
        f"## 📈 주간 거래 요약",
        f"",
        f"| 항목 | 값 |",
        f"|------|-----|",
        f"| 매수 거래 | {len(weekly_trades_buy)}건 |",
        f"| 매도 거래 | {len(weekly_trades_sell)}건 |",
        f"| 현재 보유 | {len(current_holdings)}개 종목 |",
        f"| 보유 총액 | ₩{int(total_value):,} |",
        f"",
    ]
    
    # 매수 거래
    if weekly_trades_buy:
        lines.extend([
            f"## 💰 매수 거래",
            f"",
            f"| 날짜 | Ticker | 회사명 | 가격 | 주수 |",
            f"|------|--------|--------|------|------|",
        ])
        for trade in weekly_trades_buy:
            lines.append(f"| {trade['date']} | {trade['ticker']} | {trade['company']} | ₩{int(trade['price']):,} | {int(trade['shares'])} |")
        lines.append("")
    
    # 매도 거래
    if weekly_trades_sell:
        lines.extend([
            f"## 📉 매도 거래",
            f"",
            f"| 날짜 | Ticker | 회사명 | 가격 | 주수 |",
            f"|------|--------|--------|------|------|",
        ])
        for trade in weekly_trades_sell:
            lines.append(f"| {trade['date']} | {trade['ticker']} | {trade['company']} | ₩{int(trade['price']):,} | {int(trade['shares'])} |")
        lines.append("")
    
    # 현재 보유 현황
    if current_holdings:
        lines.extend([
            f"## 📊 현재 보유 현황",
            f"",
            f"| Ticker | 주수 | 단가 | 손익 | 수익률 | 평가액 |",
            f"|--------|------|------|------|--------|--------|",
        ])
        for holding in current_holdings:
            pl_sign = "+" if holding['pl'] >= 0 else ""
            ret_sign = "+" if holding['return_rate'] >= 0 else ""
            lines.append(f"| {holding['ticker']} | {int(holding['shares'])} | ₩{int(holding['price']):,} | {pl_sign}₩{int(holding['pl']):,} | {ret_sign}{holding['return_rate']*100:.1f}% | ₩{int(holding['value']):,} |")
        lines.append("")
    
    # 손절 모니터링
    if stop_loss_data and (stop_loss_data['emergency'] or stop_loss_data['strong'] or stop_loss_data['warning']):
        lines.extend([
            f"## 🛑 손절 모니터링",
            f"",
        ])
        
        if stop_loss_data['emergency']:
            lines.extend([
                f"### 🔴 긴급 손절 (-15% 이상)",
                f"",
            ])
            for stock in stop_loss_data['emergency']:
                lines.append(f"- **{stock['ticker']}** | 진입: ₩{int(stock['entry_price']):,} | 현재: ₩{int(stock['current_price']):,} | **{stock['return_pct']}** | 손익: ₩{stock['pl']:,}")
            lines.append("")
        
        if stop_loss_data['strong']:
            lines.extend([
                f"### 🟠 강력 경고 (-10% ~ -15%)",
                f"",
            ])
            for stock in stop_loss_data['strong']:
                lines.append(f"- **{stock['ticker']}** | 진입: ₩{int(stock['entry_price']):,} | 현재: ₩{int(stock['current_price']):,} | {stock['return_pct']} | 손익: ₩{stock['pl']:,}")
            lines.append("")
        
        if stop_loss_data['warning']:
            lines.extend([
                f"### 🟡 경고 (-5% ~ -10%)",
                f"",
            ])
            for stock in stop_loss_data['warning']:
                lines.append(f"- **{stock['ticker']}** | 진입: ₩{int(stock['entry_price']):,} | 현재: ₩{int(stock['current_price']):,} | {stock['return_pct']} | 손익: ₩{stock['pl']:,}")
            lines.append("")
    
    # 다음주 매매신호
    if current_week_tickers:
        lines.extend([
            f"## 🎯 스크리너 Top 10 (오늘 기준)",
            f"",
        ])
        us_tickers = [t for t in current_week_tickers if not t[0].endswith(('.KS', '.KQ'))]
        kr_tickers = [t for t in current_week_tickers if t[0].endswith(('.KS', '.KQ'))]
        
        if us_tickers:
            lines.append(f"### 미국 종목")
            for ticker, company in us_tickers:
                lines.append(f"- **{ticker}** - {company}")
            lines.append("")
        
        if kr_tickers:
            lines.append(f"### 한국 종목")
            for ticker, company in kr_tickers:
                lines.append(f"- **{ticker}** - {company}")
            lines.append("")
    
    lines.extend([
        f"---",
        f"",
        f"**자동 생성**: Supply-Demand Dashboard 주간 분석 시스템",
    ])
    
    # 파일 저장
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    
    print(f"✅ 주간 보고서 생성 완료")
    print(f"   파일: {report_file}")
    print(f"   기간: {start_date} ~ {end_date}")
    print(f"   매수: {len(weekly_trades_buy)}건, 매도: {len(weekly_trades_sell)}건")
    print(f"   현재 보유: {len(current_holdings)}개 종목 (총액: ₩{int(total_value):,})")
    if stop_loss_data:
        print(f"   손절 대상: {stop_loss_data['total']}개")

except Exception as e:
    print(f"❌ 에러: {e}")
    import traceback
    traceback.print_exc()
