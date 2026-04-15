#!/usr/bin/env python3
"""매월 1일 10:00 - 월간 백테스트 및 성과 분석
- Portfolio Status에서 현재 포지션 수익률 계산
- track-returns 데이터와 통합해서 월간 성과 분석
"""
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os
import glob
from pathlib import Path

tracker_file = "/Users/kimdoohwan/Documents/Obsidian Vault/주식/스크리너백테스트/매매기록.xlsx"
data_dir = os.path.expanduser("~/Documents/Obsidian Vault/주식/스크리너백테스트")

try:
    wb = load_workbook(tracker_file)

    month = datetime.now().strftime("%Y-%m")
    last_month = (datetime.now().replace(day=1) - timedelta(days=1)).strftime("%Y-%m")

    # Portfolio Status에서 월간 거래 및 수익률 계산
    ps_ws = wb["Portfolio Status"]
    pt_ws = wb["Portfolio Trades"]

    total_pnl = 0.0
    total_return_pct = 0.0
    trade_count = 0
    closed_trades = []

    # Portfolio Trades에서 이번 달 거래 조회
    for row in range(2, pt_ws.max_row + 1):
        entry_date = pt_ws.cell(row=row, column=6).value  # Entry Date
        exit_price = pt_ws.cell(row=row, column=8).value  # Exit Price

        if not entry_date:
            continue

        # 문자열로 변환해서 월 확인
        entry_date_str = str(entry_date)[:7]

        # 지난 달 데이터만 처리
        if entry_date_str == last_month or entry_date_str.startswith(last_month):
            ticker = pt_ws.cell(row=row, column=2).value
            entry_price = pt_ws.cell(row=row, column=7).value
            shares = pt_ws.cell(row=row, column=5).value

            if exit_price and entry_price and shares:
                try:
                    pnl = (float(exit_price) - float(entry_price)) * float(shares)
                    total_pnl += pnl
                    return_pct = ((float(exit_price) - float(entry_price)) / float(entry_price)) * 100
                    total_return_pct += return_pct
                    trade_count += 1
                    closed_trades.append({
                        'ticker': ticker,
                        'shares': shares,
                        'entry_price': entry_price,
                        'exit_price': exit_price,
                        'pnl': pnl,
                        'return_pct': return_pct
                    })
                except (ValueError, TypeError):
                    pass

    # track-returns 데이터 통합
    returns_files = sorted(glob.glob(f"{data_dir}/*_수익률추적.md"))
    if returns_files:
        latest_returns_file = returns_files[-1]
        try:
            with open(latest_returns_file, 'r', encoding='utf-8') as f:
                content = f.read()
        except:
            pass

    # Monthly Summary 시트에 기록
    ms_ws = wb["Monthly Summary"]

    # 기본 통계
    last_row = 2
    for row in range(2, ms_ws.max_row + 1):
        if ms_ws.cell(row=row, column=1).value is None:
            last_row = row
            break

    ms_ws.cell(row=last_row, column=1).value = last_month
    ms_ws.cell(row=last_row, column=2).value = trade_count
    ms_ws.cell(row=last_row, column=3).value = round(total_pnl, 2) if trade_count > 0 else 0
    ms_ws.cell(row=last_row, column=4).value = round(total_return_pct / trade_count, 2) if trade_count > 0 else 0
    ms_ws.cell(row=last_row, column=5).value = len(closed_trades)  # 수익거래
    ms_ws.cell(row=last_row, column=6).value = trade_count - len(closed_trades)  # 손실거래

    wb.save(tracker_file)

    print(f"✅ 월간 백테스트 완료 ({last_month})")
    print(f"   거래: {trade_count}개")
    print(f"   총 손익: {total_pnl:.2f} KRW")
    print(f"   평균 수익률: {total_return_pct/trade_count:.2f}%" if trade_count > 0 else "   No trades")

except Exception as e:
    print(f"❌ 에러: {e}")
    import traceback
    traceback.print_exc()
