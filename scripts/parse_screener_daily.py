#!/usr/bin/env python3
"""매일 07:00 - 스크리너 리포트를 파싱해서 매매기록.xlsx의 Daily Screener 시트에 기록"""

import os
from datetime import datetime
from openpyxl import load_workbook
import re

today = datetime.now().strftime("%Y-%m-%d")
report_file = os.path.expanduser(f"~/Documents/Obsidian Vault/주식/스크리너백테스트/{today}/{today}_리포트.md")
tracker_file = os.path.expanduser("~/Documents/Obsidian Vault/주식/스크리너백테스트/매매기록.xlsx")

def parse_markdown_table(lines):
    """마크다운 테이블을 파싱해서 데이터 리스트 반환"""
    results = []
    in_table = False
    for line in lines:
        if '|' not in line:
            continue
        # 헤더 건너뛰기 (--- 포함된 줄)
        if '---' in line:
            in_table = True
            continue
        if not in_table:
            continue

        # 테이블 행 파싱
        parts = [p.strip() for p in line.split('|')]
        # 첫/마지막 요소는 빈 문자열 (|로 시작/끝나므로)
        if len(parts) < 6:
            continue

        parts = parts[1:-1]  # 양쪽 빈 요소 제거

        try:
            rank = int(parts[0])
            ticker = parts[1].strip('**')
            company = parts[2]
            price_str = parts[3].strip('$₩').replace(',', '')
            tpr = parts[4]
            rpr = parts[5]
            rs = parts[6] if len(parts) > 6 else '0'
            score = parts[7] if len(parts) > 7 else '0'

            results.append({
                'rank': rank,
                'ticker': ticker,
                'company': company,
                'price': float(price_str) if price_str.replace('.', '').isdigit() else 0,
                'tpr': tpr,
                'rpr': float(rpr) if rpr.replace('.', '').replace('-', '').isdigit() else 0,
                'rs': float(rs) if rs.replace('.', '').replace('-', '').isdigit() else 0,
                'score': float(score) if score.replace('.', '').replace('-', '').isdigit() else 0,
            })
        except (ValueError, IndexError):
            continue

    return results

try:
    # 리포트 파일 읽기
    if not os.path.exists(report_file):
        print(f"❌ 리포트 파일 없음: {report_file}")
        exit(1)

    with open(report_file, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')

    # US Top 10과 KR Top 10 분리
    us_section_idx = -1
    kr_section_idx = -1
    for i, line in enumerate(lines):
        if '## 미국 Top' in line:
            us_section_idx = i
        elif '## 한국 Top' in line:
            kr_section_idx = i

    if us_section_idx == -1:
        print("❌ US 섹션을 찾을 수 없음")
        exit(1)

    # 각 섹션 파싱
    us_lines = lines[us_section_idx:kr_section_idx] if kr_section_idx != -1 else lines[us_section_idx:]
    kr_lines = lines[kr_section_idx:] if kr_section_idx != -1 else []

    us_data = parse_markdown_table(us_lines)
    kr_data = parse_markdown_table(kr_lines)

    # 매매기록.xlsx 열기
    wb = load_workbook(tracker_file)
    ws = wb["Daily Screener"]

    # 기존 데이터 초기화 (헤더 제외)
    for row in range(2, ws.max_row + 1):
        for col in range(1, 13):
            ws.cell(row=row, column=col).value = None

    # 헤더 설정 (첫 줄)
    headers = ['Date', 'Region', 'Rank', 'Ticker', 'Company', 'Price', 'TPR', 'RPR', 'RS', 'Score', 'Signal', 'Notes']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header

    # US Top 20 기록
    row = 2
    for stock in us_data:
        ws.cell(row=row, column=1).value = today
        ws.cell(row=row, column=2).value = 'US'
        ws.cell(row=row, column=3).value = stock['rank']
        ws.cell(row=row, column=4).value = stock['ticker']
        ws.cell(row=row, column=5).value = stock['company']
        ws.cell(row=row, column=6).value = stock['price']
        ws.cell(row=row, column=7).value = stock['tpr']
        ws.cell(row=row, column=8).value = stock['rpr']
        ws.cell(row=row, column=9).value = stock['rs']
        ws.cell(row=row, column=10).value = stock['score']
        row += 1

    # KR Top 20 기록
    for stock in kr_data:
        ws.cell(row=row, column=1).value = today
        ws.cell(row=row, column=2).value = 'KR'
        ws.cell(row=row, column=3).value = stock['rank']
        ws.cell(row=row, column=4).value = stock['ticker']
        ws.cell(row=row, column=5).value = stock['company']
        ws.cell(row=row, column=6).value = stock['price']
        ws.cell(row=row, column=7).value = stock['tpr']
        ws.cell(row=row, column=8).value = stock['rpr']
        ws.cell(row=row, column=9).value = stock['rs']
        ws.cell(row=row, column=10).value = stock['score']
        row += 1

    # 저장
    wb.save(tracker_file)
    print(f"✅ {today} 스크리너 데이터 기록 완료")
    print(f"   US Top 10: {len(us_data)}개, KR Top 10: {len(kr_data)}개")

except Exception as e:
    print(f"❌ 에러: {e}")
    import traceback
    traceback.print_exc()
