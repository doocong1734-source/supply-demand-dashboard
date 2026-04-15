#!/usr/bin/env python3
"""매주 월요일 09:00 - 주간 리밸런싱 신호 생성
- Daily Screener에서 최신 Top 20 조회
- Portfolio Status와 비교해서 리밸런스 신호 생성
- Weekly Summary 시트에 기록
"""
from datetime import datetime, timedelta
from openpyxl import load_workbook

tracker_file = "/Users/kimdoohwan/Documents/Obsidian Vault/주식/스크리너백테스트/매매기록.xlsx"

try:
    wb = load_workbook(tracker_file)

    # Daily Screener 시트에서 최신 Top 20 조회
    ds_ws = wb["Daily Screener"]

    # 가장 최근 날짜의 Top 20 찾기
    latest_date = None
    top_20_tickers = set()

    for row in range(2, ds_ws.max_row + 1):
        date_val = ds_ws.cell(row=row, column=1).value
        ticker = ds_ws.cell(row=row, column=4).value
        rank = ds_ws.cell(row=row, column=3).value

        if date_val and ticker and rank and rank <= 20:
            if latest_date is None or str(date_val) > str(latest_date):
                latest_date = date_val
                top_20_tickers.clear()

            if str(date_val) == str(latest_date):
                top_20_tickers.add(ticker)

    # Portfolio Status에서 현재 보유 종목 조회
    ps_ws = wb["Portfolio Status"]
    held_tickers = {}

    for row in range(2, ps_ws.max_row + 1):
        ticker = ps_ws.cell(row=row, column=2).value
        shares = ps_ws.cell(row=row, column=3).value

        if ticker and shares:
            try:
                shares_num = float(shares) if isinstance(shares, str) else shares
                if shares_num > 0:
                    held_tickers[ticker] = shares_num
            except (ValueError, TypeError):
                pass

    # Weekly Summary 시트에 신호 기록
    ws_week = wb["Weekly Summary"]

    today = datetime.now()
    week_num = today.isocalendar()[1]

    # 리밸런싱 신호 계산
    buy_signals = top_20_tickers - set(held_tickers.keys())
    sell_signals = set(held_tickers.keys()) - top_20_tickers

    # Weekly Summary 마지막 행에 기록
    last_row = ws_week.max_row + 1

    ws_week.cell(row=last_row, column=1).value = today.strftime("%Y-%m-%d")
    ws_week.cell(row=last_row, column=2).value = f"Week {week_num}"
    ws_week.cell(row=last_row, column=3).value = f"Top 10: {', '.join(sorted(top_20_tickers)[:5])}..." if top_20_tickers else "No data"
    ws_week.cell(row=last_row, column=4).value = f"Buy: {', '.join(sorted(buy_signals))}" if buy_signals else "None"
    ws_week.cell(row=last_row, column=5).value = f"Sell: {', '.join(sorted(sell_signals))}" if sell_signals else "None"

    wb.save(tracker_file)

    print(f"✅ 주간 리밸런싱 신호 생성 완료")
    print(f"   기준일: {latest_date}")
    print(f"   Top 20: {len(top_20_tickers)}개")
    print(f"   현재보유: {len(held_tickers)}개")
    print(f"   매수신호: {len(buy_signals)}개")
    print(f"   매도신호: {len(sell_signals)}개")

except Exception as e:
    print(f"❌ 에러: {e}")
    import traceback
    traceback.print_exc()
