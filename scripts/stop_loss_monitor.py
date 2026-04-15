#!/usr/bin/env python3
"""손절 모니터링 - Portfolio Status에서 손실 종목 감지
손절 기준: -5% (경고), -10% (강력 경고), -15% (긴급)
"""

from openpyxl import load_workbook
import json
from datetime import datetime

tracker_file = "/Users/kimdoohwan/Documents/Obsidian Vault/주식/스크리너백테스트/매매기록.xlsx"

# 손절 기준
STOP_LOSS_LEVELS = {
    'warning': -0.05,      # -5%
    'strong': -0.10,       # -10%
    'emergency': -0.15     # -15%
}

try:
    wb = load_workbook(tracker_file)
    
    # Portfolio Status 시트에서 손실 종목 조회
    status_ws = wb["Portfolio Status"]
    
    warning_stocks = []      # -5% ~ -10%
    strong_stocks = []       # -10% ~ -15%
    emergency_stocks = []    # -15% 이상
    
    for row in range(2, status_ws.max_row + 1):
        ticker = status_ws.cell(row=row, column=2).value
        shares = status_ws.cell(row=row, column=3).value
        price = status_ws.cell(row=row, column=4).value
        entry_price = status_ws.cell(row=row, column=5).value
        pl = status_ws.cell(row=row, column=6).value
        
        if not ticker or not shares:
            continue
        
        try:
            shares_num = float(shares) if isinstance(shares, str) else shares
            if shares_num <= 0:
                continue
            
            price_num = float(price) if isinstance(price, str) else (price or 0)
            entry_num = float(entry_price) if isinstance(entry_price, str) else (entry_price or 0)
            pl_num = float(pl) if isinstance(pl, str) else (pl or 0)
            
            # 수익률 계산
            if entry_num > 0:
                return_rate = (price_num - entry_num) / entry_num
            else:
                continue
            
            # 손실 종목만 분류
            if return_rate < 0:
                stock_info = {
                    'ticker': ticker,
                    'shares': int(shares_num),
                    'entry_price': entry_num,
                    'current_price': price_num,
                    'pl': int(pl_num),
                    'return_rate': return_rate,
                    'return_pct': f"{return_rate*100:.1f}%"
                }
                
                if return_rate <= STOP_LOSS_LEVELS['emergency']:
                    emergency_stocks.append(stock_info)
                elif return_rate <= STOP_LOSS_LEVELS['strong']:
                    strong_stocks.append(stock_info)
                elif return_rate <= STOP_LOSS_LEVELS['warning']:
                    warning_stocks.append(stock_info)
        
        except (ValueError, TypeError):
            continue
    
    # 손실률로 정렬
    warning_stocks.sort(key=lambda x: x['return_rate'])
    strong_stocks.sort(key=lambda x: x['return_rate'])
    emergency_stocks.sort(key=lambda x: x['return_rate'])
    
    # 결과 출력
    print("=" * 70)
    print("🛑 손절 모니터링 리포트")
    print("=" * 70)
    print(f"\n📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    total_loss_stocks = len(warning_stocks) + len(strong_stocks) + len(emergency_stocks)
    
    if total_loss_stocks == 0:
        print("✅ 손절 대상 종목 없음 (모두 수익 또는 보유 안 함)")
    else:
        print(f"⚠️  손절 대상 종목: {total_loss_stocks}개\n")
        
        # 긴급 손절 (-15% 이상)
        if emergency_stocks:
            print("🔴 긴급 손절 (-15% 이상)")
            print("-" * 70)
            for stock in emergency_stocks:
                print(f"  {stock['ticker']:15} | 진입: ₩{int(stock['entry_price']):>8,} | 현재: ₩{int(stock['current_price']):>8,} | {stock['return_pct']:>7} | 손익: {stock['pl']:>10,}원")
            print()
        
        # 강력 경고 (-10% ~ -15%)
        if strong_stocks:
            print("🟠 강력 경고 (-10% ~ -15%)")
            print("-" * 70)
            for stock in strong_stocks:
                print(f"  {stock['ticker']:15} | 진입: ₩{int(stock['entry_price']):>8,} | 현재: ₩{int(stock['current_price']):>8,} | {stock['return_pct']:>7} | 손익: {stock['pl']:>10,}원")
            print()
        
        # 경고 (-5% ~ -10%)
        if warning_stocks:
            print("🟡 경고 (-5% ~ -10%)")
            print("-" * 70)
            for stock in warning_stocks:
                print(f"  {stock['ticker']:15} | 진입: ₩{int(stock['entry_price']):>8,} | 현재: ₩{int(stock['current_price']):>8,} | {stock['return_pct']:>7} | 손익: {stock['pl']:>10,}원")
            print()
    
    # JSON으로 저장 (주간 리포트에서 사용)
    stop_loss_data = {
        'timestamp': datetime.now().isoformat(),
        'emergency': emergency_stocks,
        'strong': strong_stocks,
        'warning': warning_stocks,
        'total': total_loss_stocks
    }
    
    with open('/tmp/stop_loss_data.json', 'w', encoding='utf-8') as f:
        json.dump(stop_loss_data, f, ensure_ascii=False, indent=2)
    
    print("\n" + "=" * 70)
    if total_loss_stocks == 0:
        print("✅ 모니터링 완료 - 손절 대상 없음")
    else:
        print(f"⚠️  모니터링 완료 - {total_loss_stocks}개 종목 관찰 중")
        if emergency_stocks:
            print(f"🔴 긴급: {len(emergency_stocks)}개 | 🟠 강경: {len(strong_stocks)}개 | 🟡 경고: {len(warning_stocks)}개")
    print("=" * 70)

except Exception as e:
    print(f"❌ 에러: {e}")
    import traceback
    traceback.print_exc()
